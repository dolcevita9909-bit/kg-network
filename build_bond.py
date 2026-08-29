#!/usr/bin/env python3
"""project Argus 시트 -> 채권시장 지식그래프 (bond/index.html) 생성."""
import json, sys, urllib.request
from openpyxl import load_workbook

SHEET_ID = "1xPIto-HKznjkWgZ0NCtLyIENkB2Bok0i7Iz1pYCou4A"
URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
NODE_SHEET = "01_Node_Master"
PAT_SHEET = "04_Pattern_Master"


def s(v):
    return "" if v is None else str(v).strip()


def rows_of(ws):
    return [[s(c) for c in r] for r in ws.iter_rows(values_only=True)]


def main():
    urllib.request.urlretrieve(URL, "argus.xlsx")
    wb = load_workbook("argus.xlsx", read_only=True, data_only=True)

    nr = rows_of(wb[NODE_SHEET])
    hdr = [i for i, r in enumerate(nr) if r and r[0] == "Node_ID"]
    if not hdr:
        sys.exit("Node_ID 헤더를 찾지 못함")
    bounds = hdr + [len(nr)]

    nodes = {}
    for bi, start in enumerate(hdr):
        end = bounds[bi + 1]
        head = nr[start]
        wide = "Upstream_Nodes" in head          # 두 번째 블록(18열) 여부
        imp_i = head.index("Importance") if "Importance" in head else -1
        for r in nr[start + 1:end]:
            if not r or not r[0] or r[0] == "Node_ID":
                continue
            g = lambda i: r[i] if i < len(r) else ""
            nodes[r[0]] = {
                "id": r[0], "name": g(1), "ont": g(3), "dom": g(4),
                "par": [x.strip() for x in g(6).split(";") if x.strip() and x.strip() != "-"],
                "q": g(7), "def": g(8), "th": g(9),
                "up": g(10) if wide else "", "dn": g(11) if wide else "",
                "ld": g(12) if wide else "", "lg": g(13) if wide else "",
                "etf": g(14) if wide else "",
                "imp": g(imp_i) if imp_i >= 0 else "",
            }

    # 시트에서 참조되지만 정의가 없는 노드 보완
    if "INF0000" not in nodes:
        nodes["INF0000"] = {"id": "INF0000", "name": "물가", "ont": "Concept", "dom": "Inflation",
                            "par": [], "q": "물가란 무엇인가?",
                            "def": "재화와 서비스의 전반적인 가격 수준",
                            "th": "물가는 금리·정책·기업이익을 움직이는 매크로의 출발점이다.",
                            "up": "", "dn": "CPI, PCE, 근원물가", "ld": "", "lg": "", "etf": "",
                            "imp": "★★★★★"}
    if "INT0700" not in nodes:
        nodes["INT0700"] = {"id": "INT0700", "name": "기간프리미엄", "ont": "Derived", "dom": "Interest",
                            "par": ["INT0200"], "q": "기간프리미엄이란 무엇인가?",
                            "def": "장기채 보유에 요구되는 추가 보상",
                            "th": "기간프리미엄은 재정위험과 인플레 불확실성을 장기금리에 반영하는 통로이다.",
                            "up": "재정적자, 인플레이션 불확실성, 국채수급",
                            "dn": "장기금리, 수익률곡선",
                            "ld": "ACM Term Premium, 국채발행계획", "lg": "10년물 금리",
                            "etf": "장기국채", "imp": "★★★★★"}

    # 계층 깊이
    def depth(nid, seen=None):
        seen = seen or set()
        if nid in seen:
            return 0
        seen.add(nid)
        n = nodes.get(nid)
        if not n or not n["par"]:
            return 0
        ds = [depth(p, set(seen)) + 1 for p in n["par"] if p in nodes]
        return max(ds) if ds else 0

    for n in nodes.values():
        n["d"] = min(depth(n["id"]), 5)

    # 이름 -> ID 색인 (괄호 앞 표기도 인식)
    n2i = {}
    for n in nodes.values():
        nm = n["name"].strip()
        if nm:
            n2i[nm] = n["id"]
            if "(" in nm:
                n2i[nm.split("(")[0].strip()] = n["id"]

    links, seen = [], set()

    def add(a, b, k):
        if not a or not b or a == b:
            return
        key = f"{a}>{b}{k}"
        if key in seen:
            return
        seen.add(key)
        links.append([a, b, k])

    for n in nodes.values():
        for p in n["par"]:
            if p in nodes:
                add(p, n["id"], "H")
    for n in nodes.values():
        for x in [y.strip() for y in n["up"].split(",") if y.strip()]:
            if x in n2i:
                add(n2i[x], n["id"], "F")
        for x in [y.strip() for y in n["dn"].split(",") if y.strip()]:
            if x in n2i:
                add(n["id"], n2i[x], "F")

    # 패턴
    pr = rows_of(wb[PAT_SHEET])
    pats = []
    for r in pr[1:]:
        if not r or not r[0] or r[0] == "Pattern_ID" or len(r) < 3:
            continue
        if r[1] not in nodes:
            continue
        g = lambda i: r[i] if i < len(r) else ""
        p = {"n": r[1], "nm": g(2), "i": g(3), "m": g(4), "o": g(5), "d": g(6),
             "b": g(7), "s": g(8), "l": g(9), "c": g(10), "x": g(11)}
        pats.append({k: v for k, v in p.items() if v})

    data = {
        "nodes": [{k: v for k, v in n.items() if v not in ("", [], None)} for n in nodes.values()],
        "links": links,
        "pats": pats,
    }
    if len(data["nodes"]) < 30 or len(links) < 30:
        sys.exit(f"데이터 부족: nodes={len(data['nodes'])} links={len(links)}")

    tpl = open("template_bond.html", encoding="utf-8").read()
    html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    import os
    os.makedirs("bond", exist_ok=True)
    with open("bond/index.html", "w", encoding="utf-8") as f:
        f.write(html)
    print(f"OK nodes={len(data['nodes'])} links={len(links)} pats={len(pats)}")


if __name__ == "__main__":
    main()
