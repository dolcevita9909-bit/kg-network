#!/usr/bin/env python3
"""구글시트(xlsx export)에서 지식그래프 데이터를 추출해 index.html 생성."""
import json, sys, urllib.request
from collections import Counter
from openpyxl import load_workbook

SHEET_ID = "1wi6N5ACq7AH0_VwyfxwjSbnYypcSEZPuKXED0X21HJw"
URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

def cell(row, i):
    v = row[i] if i < len(row) else None
    return "" if v is None else str(v).strip()

def main():
    urllib.request.urlretrieve(URL, "sheet.xlsx")
    wb = load_workbook("sheet.xlsx", read_only=True, data_only=True)
    rows = {"STK": [], "THM": [], "STM": [], "STR": []}
    for ws in wb.worksheets:
        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            first = str(r[0])
            for p in rows:
                if first.startswith(p + "_"):
                    rows[p].append([("" if c is None else str(c).strip()) for c in r])
    lvl = {"1": "System", "2": "Theme", "3": "SubTheme", "4": "Detail",
           "Theme": "Theme", "SubTheme": "SubTheme", "System": "System", "Detail": "Detail"}
    nodes, links = {}, []
    for t in rows["THM"]:
        nodes[t[0]] = {"id": t[0], "name": cell(t, 1), "en": cell(t, 2), "type": "theme",
                       "level": lvl.get(cell(t, 3), "Theme"), "def": cell(t, 6)}
    for s in rows["STK"]:
        nodes[s[0]] = {"id": s[0], "name": cell(s, 2), "en": cell(s, 1), "type": "stock",
                       "level": "Stock", "system": cell(s, 6), "funds": cell(s, 4),
                       "imp": cell(s, 5), "def": cell(s, 11)}
    for m in rows["STM"]:
        sid, tid = cell(m, 1), cell(m, 3)
        if not sid or not tid:
            continue
        if tid not in nodes:
            nodes[tid] = {"id": tid, "name": cell(m, 4), "en": "", "type": "theme",
                          "level": "SubTheme", "def": "(Theme_Master 미정의)"}
        if sid in nodes:
            links.append({"s": sid, "t": tid, "role": cell(m, 6),
                          "score": cell(m, 7), "why": cell(m, 12)})
    for s in rows["STR"]:
        a, b = cell(s, 2), cell(s, 4)
        if a in nodes and b in nodes:
            links.append({"s": a, "t": b, "role": cell(s, 6),
                          "score": cell(s, 8), "why": cell(s, 13)})
    deg = Counter()
    for l in links:
        deg[l["s"]] += 1; deg[l["t"]] += 1
    for n in nodes.values():
        n["deg"] = deg[n["id"]]
    data = {"nodes": list(nodes.values()), "links": links}
    if len(nodes) < 100 or len(links) < 100:
        sys.exit(f"데이터가 비정상적으로 적음: nodes={len(nodes)} links={len(links)}")
    tpl = open("template.html", encoding="utf-8").read()
    html = tpl.replace("__DATA__", json.dumps(data, ensure_ascii=False))
    with open("index.html", "w", encoding="utf-8") as f:
        f.write(html)
    print(f"OK nodes={len(nodes)} links={len(links)}")

if __name__ == "__main__":
    main()
